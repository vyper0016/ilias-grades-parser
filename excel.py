import openpyxl
import re
from configparser import ConfigParser

import openpyxl.worksheet
from misc import url_builder
from openpyxl.formatting.rule import ColorScaleRule

INVALID_TITLE_REGEX = re.compile(r'[\\*?:/\[\]]') 

NAME_EQUIVALENTS = {
    'grade_titles': 'title',
    'grade_deadlines': 'deadline',
    'grade_submitted': 'submitted',
    'grade_value': 'grade',
    'grade_max': 'max_grade',
}

def get_sheet_list(file_path):
    wb = openpyxl.load_workbook(file_path)
    return wb.sheetnames

def cell_range_valid(s: str):
    return re.match(r'^[A-Z]+[0-9]+:[A-Z]+[0-9]+$', s)

def cell_range_to_list(s: str):
    if not cell_range_valid(s):
        raise ValueError('Invalid cell range')
    start, end = s.split(':')
    start_col = ord(start[0]) - ord('A')
    start_row = int(start[1:]) - 1
    end_col = ord(end[0]) - ord('A')
    end_row = int(end[1:]) - 1
    return [(chr(i + ord('A')) + str(j + 1)) for i in range(start_col, end_col + 1) for j in range(start_row, end_row + 1)]

def template_sheet(wb: openpyxl.Workbook, new_sheet_name, course_title, number_tests: int):
    FONT_SIZE = 14
    COLUMN_WIDTH = 13.5   

    last_test_row = 2 + number_tests
    cells_dict = {
            "title": f"A3:A{last_test_row}",
            "deadline": f"B3:B{last_test_row}",
            "submitted": f"C3:C{last_test_row}",
            "grade": f"D3:D{last_test_row}",
            "max_grade": f"E3:E{last_test_row}"
        }
    
    
    wb.create_sheet(new_sheet_name)
    ws = wb[new_sheet_name]        

    for cell in cell_range_to_list(f'A1:F{last_test_row + 2}'):
        ws[cell].font = openpyxl.styles.Font(size=FONT_SIZE)
    
    ws.row_dimensions[1].height = 36
        
    for key in 'ABCDEF':
        ws.column_dimensions[key].width = COLUMN_WIDTH

    ws.merge_cells('A1:F1')
    ws['A1'] = course_title
    
    columns = ['title', 'deadline', 'submitted', 'grade', 'max', 'percentage']
    
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=FONT_SIZE)
    ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    ws.column_dimensions['B'].width = 22.5
    
    for name, cell in zip(columns, cell_range_to_list('A2:F2')):
        ws[cell] = name
        ws[cell].font = openpyxl.styles.Font(bold=True, size=FONT_SIZE)
        ws[cell].alignment = openpyxl.styles.Alignment(horizontal='center')
        
    for cell in cell_range_to_list(cells_dict['deadline']):
        ws[cell].number_format = 'dd. mmm yyyy, hh:mm'
        
    for cell in cell_range_to_list(f'F3:F{last_test_row}'):
        ws[cell].number_format = '0.0%'
        ws[cell].value = f'=IF(E{cell[1]}<>0,D{cell[1]}/E{cell[1]},0)'
    
    for i in [1, 2]:
        ws.merge_cells(f'A{last_test_row + i}:E{last_test_row + i}')
        ws[f'A{last_test_row + i}'].font = openpyxl.styles.Font(bold=True, size=FONT_SIZE)
        ws[f'A{last_test_row + i}'].alignment = openpyxl.styles.Alignment(horizontal='center')
        ws[f'F{last_test_row + i}'].number_format = '0.00%'
    
    ws[f'A{last_test_row + 1}'] = 'Average'  
    ws[f'F{last_test_row + 1}'] = f'=AVERAGE(F3:F{last_test_row})'
    
    ws[f'A{last_test_row + 2}'] = 'Zulassung'  
    ws[f'F{last_test_row + 2}'] = f'=F{last_test_row + 1}*2'
    
    rule = ColorScaleRule(start_type='percentile', start_value=0, start_color='f8696b',
                            mid_type='percentile', mid_value=50, mid_color='b1d580',
                            end_type='percentile', end_value=100, end_color='63be7b')
    ws.conditional_formatting.add(f'F3:F{last_test_row + 2}', rule)
    
    return cells_dict

def sheet_title_valid(sheet_title: str, verbose=False):
    m = INVALID_TITLE_REGEX.search(sheet_title)
    if m:
        if verbose:
            msg = "Invalid character '{0}' found in sheet title".format(m.group(0))
            print(msg)
        return False   
    return True 

def validate_sheet_title(sheet_title: str):
    sheet_title = re.sub(INVALID_TITLE_REGEX, '', sheet_title)
    return sheet_title

def make_from_template(file_path, course_grades: dict, course_excel_dict: dict):
    new_sheet_name = course_excel_dict['sheet_name']
    number_tests = course_excel_dict['number_tests']
    wb = openpyxl.load_workbook(file_path)
    
    if new_sheet_name in wb.sheetnames:
        raise ValueError('Sheet name already exists in workbook')
    
    if 'todelete' in wb.sheetnames:
        del wb['todelete']
    
    course_title = course_grades['title']
    cells_dict = template_sheet(wb, new_sheet_name, course_title, number_tests)
    save_wb_retriable(wb, file_path)
    course_excel_dict['cells'] = cells_dict
    save_to_excel(file_path, course_grades, course_excel_dict)
    return course_excel_dict
    
def save_to_excel(file_path, course_grades: dict, course_excel_dict: dict):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[course_excel_dict['sheet_name']]
    for attr in course_excel_dict['cells']:
        for grade, cell in zip(course_grades['grades'], cell_range_to_list(course_excel_dict['cells'][attr])):
            if attr not in grade:
                continue
            if attr in  ['submitted', 'grade'] and ws[cell].value is not None:
                continue
            ws[cell] = grade[attr]
            if attr == 'title' and 'url' in grade and grade['url'] is not None:
                ws[cell].hyperlink = url_builder(grade['url'])
    
    save_wb_retriable(wb, file_path)  

def save_wb_retriable(wb, file_path):
    while True:
        try:
            wb.save(file_path)
            break
        except PermissionError:
            print(f"Could not save to {file_path}. Please close the file and press enter to retry.")
            input()
        except Exception as e:
            print(f"An error occurred: {e}. Please close the file and press enter to retry.")
            print(e.with_traceback())
            return
    
if __name__ == '__main__':
    quit()
    import json
    with open('data/grades.json') as f:
        grades = json.load(f)
    make_from_template(r"D:\sciebo\hhu\test.xlsx", grades['1639723'], 'test')
    